from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
PDF_A = FIXTURES / "layout_a_classic_invoice.pdf"


def test_healthz(client):
    assert client.get("/healthz").json() == {"status": "ok"}


def test_extract_endpoint_end_to_end(client, store):
    resp = client.post(
        "/extract",
        files={"file": ("layout_a_classic_invoice.pdf", PDF_A.read_bytes(), "application/pdf")},
    )
    assert resp.status_code == 201
    record = resp.json()
    assert record["disposition"] == "approved"
    assert record["payload"]["vendor_name"] == "Meridian Office Supplies Ltd."
    assert abs(record["payload"]["total_amount"] - 2148.84) < 0.01
    assert record["findings"] == []

    listed = client.get("/api/documents").json()
    assert len(listed) == 1
    assert listed[0]["id"] == record["id"]


def test_extract_rejects_text_file(client):
    resp = client.post(
        "/extract",
        files={"file": ("notes.txt", b"vendor: acme total: 5", "text/plain")},
    )
    assert resp.status_code == 422
    assert "unsupported file type" in resp.json()["detail"]


def test_extract_empty_upload(client):
    resp = client.post("/extract", files={"file": ("empty.pdf", b"", "application/pdf")})
    assert resp.status_code == 400


def test_excel_export_contains_clean_and_flagged(client, store):
    client.post("/extract", files={"file": ("a.pdf", PDF_A.read_bytes(), "application/pdf")})
    xlsx = client.get("/export/excel.xlsx")
    assert xlsx.status_code == 200
    assert xlsx.content.startswith(b"PK")

    import io

    wb = load_workbook(io.BytesIO(xlsx.content))
    summary = wb["Summary"]
    row = [c.value for c in summary[2]]
    assert row[3] == "Meridian Office Supplies Ltd."
    assert wb["LineItems"].max_row == 4


def test_audit_log_jsonl_download(client, store):
    resp = client.post("/extract", files={"file": ("a.pdf", PDF_A.read_bytes(), "application/pdf")})
    doc_id = resp.json()["id"]
    audit = client.get("/export/audit-log.jsonl")
    assert audit.status_code == 200
    lines = [line for line in audit.text.strip().splitlines() if line]
    assert len(lines) == 1
    assert doc_id in lines[0]


def test_review_flow_via_api(client, store):
    resp = client.post("/extract", files={"file": ("a.pdf", PDF_A.read_bytes(), "application/pdf")})
    doc_id = resp.json()["id"]
    patched = client.patch(
        f"/api/documents/{doc_id}/review",
        json={"review_status": "approved", "review_note": "spot checked"},
    )
    assert patched.status_code == 200
    body = patched.json()
    assert body["review_status"] == "approved"

    bad = client.patch(f"/api/documents/{doc_id}/review", json={"review_status": "maybe"})
    assert bad.status_code == 400


def test_dashboard_and_detail_pages_render(client, store):
    resp = client.post("/extract", files={"file": ("a.pdf", PDF_A.read_bytes(), "application/pdf")})
    doc_id = resp.json()["id"]
    dash = client.get("/", headers={"accept": "text/html"})
    assert dash.status_code == 200
    assert "LedgerLens" in dash.text

    detail = client.get(f"/documents/{doc_id}", headers={"accept": "text/html"})
    assert detail.status_code == 200
    assert "Meridian" in detail.text

    missing = client.get("/documents/nonexistent")
    assert missing.status_code == 404
