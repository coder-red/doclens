from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.exports import build_workbook
from app.models import Finding
from app.store import Store

from tests.conftest import FakeProvider, _default_payload

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture()
def populated_store(tmp_path) -> Store:
    from app.pipeline import process_document
    from app.config import Settings

    store = Store(tmp_path / "s.db")
    settings = Settings(db_path=tmp_path / "s.db")

    clean_payload = _default_payload()
    provider_clean = FakeProvider([clean_payload])
    data = (ROOT / "fixtures" / "layout_a_classic_invoice.pdf").read_bytes()
    process_document(data, "clean.pdf", provider=provider_clean, store=store, settings=settings)

    flagged_payload = _default_payload()
    flagged_payload["currency"] = None
    provider_flagged = FakeProvider([flagged_payload])
    process_document(data, "flagged.pdf", provider=provider_flagged, store=store, settings=settings)

    provider_garbage = FakeProvider([{"nonsense": True}])
    process_document(data, "garbage.pdf", provider=provider_garbage, store=store, settings=settings)
    return store


def test_insert_and_list_roundtrip(populated_store):
    rows = populated_store.list_documents()
    assert len(rows) == 3
    dispositions = {r["disposition"] for r in rows}
    assert dispositions == {"approved", "flagged", "failed"}


def test_filter_by_disposition(populated_store):
    approved = populated_store.list_documents(disposition="approved")
    assert len(approved) == 1
    assert approved[0]["payload"]["vendor_name"] == "Meridian Office Supplies Ltd."


def test_get_document_includes_raw_response(populated_store):
    approved_id = populated_store.list_documents(disposition="approved")[0]["id"]
    record = populated_store.get_document(approved_id)
    assert record["raw_response"]["document_type"] == "invoice"
    assert record["content_sha256"]


def test_update_review(populated_store):
    doc_id = populated_store.list_documents(disposition="flagged")[0]["id"]
    assert populated_store.update_review(doc_id, "approved", "verified against original")
    record = populated_store.get_document(doc_id)
    assert record["review_status"] == "approved"
    assert record["review_note"] == "verified against original"


def test_counts_by_disposition(populated_store):
    counts = populated_store.counts_by_disposition()
    assert counts == {"approved": 1, "flagged": 1, "failed": 1}


def test_excel_workbook_structure(populated_store):
    records = populated_store.list_documents(limit=100)
    content = build_workbook(records)
    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["Summary", "LineItems"]
    summary = wb["Summary"]
    assert summary.max_row == 4
    headers = [c.value for c in summary[1]]
    assert "flags" in headers and "total_amount" in headers
    line_items = wb["LineItems"]
    assert line_items.max_row == 7


def test_audit_log_lines_are_jsonl(populated_store):
    lines = populated_store.audit_log_lines()
    assert len(lines) == 3
    parsed = [json.loads(line) for line in lines]
    by_file = {p["source_filename"]: p for p in parsed}
    flagged = by_file["flagged.pdf"]
    assert flagged["disposition"] == "flagged"
    assert any(f["rule_id"] == "V004" for f in flagged["findings"])

