from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_SUMMARY_HEADERS = [
    "document_id", "processed_at", "source_file", "vendor", "invoice_number",
    "document_type", "document_date", "due_date", "currency", "subtotal",
    "tax_amount", "total_amount", "line_item_count", "provider", "model",
    "flags", "review_status",
]
_LINE_HEADERS = [
    "document_id", "vendor", "invoice_number", "row", "description",
    "quantity", "unit_price", "line_total", "currency", "disposition",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_DISPOSITION_FILLS = {
    "approved": PatternFill("solid", fgColor="D1FAE5"),
    "flagged": PatternFill("solid", fgColor="FEF3C7"),
    "failed": PatternFill("solid", fgColor="FEE2E2"),
}


def build_workbook(records: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    _write_header(summary, _SUMMARY_HEADERS)

    line_items = wb.create_sheet("LineItems")
    _write_header(line_items, _LINE_HEADERS)

    for record in records:
        payload = record.get("payload") or {}
        disposition = record.get("disposition", "")
        row = [
            record["id"],
            record.get("created_at"),
            record.get("source_filename"),
            payload.get("vendor_name"),
            payload.get("invoice_number"),
            payload.get("document_type"),
            payload.get("document_date"),
            payload.get("due_date"),
            payload.get("currency"),
            payload.get("subtotal"),
            payload.get("tax_amount"),
            payload.get("total_amount"),
            len(payload.get("line_items") or []),
            record.get("provider"),
            record.get("model"),
            "; ".join(f"{f['rule_id']}:{f['field']}" for f in record.get("findings", [])),
            record.get("review_status"),
        ]
        summary.append(row)
        fill = _DISPOSITION_FILLS.get(disposition)
        if fill:
            for col in range(1, len(_SUMMARY_HEADERS) + 1):
                summary.cell(row=summary.max_row, column=col).fill = fill
        if record.get("findings"):
            flags_cell = summary.cell(
                row=summary.max_row, column=_SUMMARY_HEADERS.index("flags") + 1
            )
            flags_cell.font = Font(color="B45309")

        for idx, item in enumerate(payload.get("line_items") or [], start=1):
            line_items.append([
                record["id"],
                payload.get("vendor_name"),
                payload.get("invoice_number"),
                idx,
                item.get("description"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("line_total"),
                payload.get("currency"),
                disposition,
            ])

    _autosize(summary)
    _autosize(line_items)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def _autosize(sheet) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = max(
            len(str(sheet.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(sheet.max_row, 200) + 1)
        )
        sheet.column_dimensions[letter].width = min(max(widest + 2, 10), 48)
